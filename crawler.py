"""
携程旅游产品爬虫核心逻辑
"""

import time
import random
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from urllib.parse import quote
import logging
import re

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class CtripTourCrawler:
    def __init__(self, destination="武汉", output_file="tours.xlsx"):
        self.destination = destination
        self.output_file = output_file
        self.driver = self._init_driver()
        self.tours = []
        self.session = requests.Session()
        self.session.headers.update(self._get_headers())
        
    def _init_driver(self):
        """初始化 Selenium WebDriver"""
        options = Options()
        # 可选：添加 --headless 参数在后台运行
        # options.add_argument("--headless")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36")
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option('useAutomationExtension', False)
        
        try:
            driver = webdriver.Chrome(options=options)
            driver.set_page_load_timeout(30)
            logger.info("✅ Chrome 浏览器已启动")
            return driver
        except Exception as e:
            logger.error(f"❌ 启动 Chrome 失败: {str(e)}")
            logger.info("请确保已安装 Chrome 浏览器和 ChromeDriver")
            raise
    
    @staticmethod
    def _get_headers():
        """获取随机 User-Agent"""
        user_agents = [
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:121.0) Gecko/20100101 Firefox/121.0",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36 Edg/120.0.0.0",
        ]
        return {
            "User-Agent": random.choice(user_agents),
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
            "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
            "Accept-Encoding": "gzip, deflate",
            "DNT": "1",
            "Connection": "keep-alive",
            "Upgrade-Insecure-Requests": "1",
            "Referer": "https://www.ctrip.com/",
        }
    
    def _random_delay(self, min_sec=1, max_sec=3):
        """随机延迟"""
        delay = random.uniform(min_sec, max_sec)
        time.sleep(delay)
    
    def crawl_all_tours(self):
        """爬取所有旅游产品"""
        tours = []
        page = 1
        max_pages = 100  # 防止无限循环
        consecutive_empty = 0  # 连续空页数
        
        while page <= max_pages:
            logger.info(f"正在爬取第 {page} 页...")
            
            try:
                page_tours = self._crawl_page(page)
                
                if not page_tours:
                    consecutive_empty += 1
                    logger.warning(f"第 {page} 页没有产品 (连续空页: {consecutive_empty})")
                    
                    # 如果连续3页都没有产品，则停止
                    if consecutive_empty >= 3:
                        logger.info("连续3页未找到产品，停止爬取")
                        break
                else:
                    consecutive_empty = 0  # 重置计数
                    tours.extend(page_tours)
                    logger.info(f"第 {page} 页爬取了 {len(page_tours)} 个产品，累计 {len(tours)} 个")
                
                page += 1
                self._random_delay(2, 5)  # 页面间延迟
                
            except Exception as e:
                logger.error(f"爬取第 {page} 页时出错: {str(e)}")
                page += 1
                self._random_delay(3, 6)
        
        logger.info(f"✅ 爬取完成，总共获得 {len(tours)} 个产品")
        return tours
    
    def _crawl_page(self, page):
        """爬取单个页面"""
        try:
            # 构造 URL - 携程跟团游产品搜索页
            url = "https://you.ctrip.com/tours/search/vacations"
            
            # 参数
            params = {
                "startingcityid": "2",  # 出发地参数
                "keyword": self.destination,  # 目的地关键词
                "page": page,
            }
            
            # 拼接完整 URL
            query_string = "&".join([f"{k}={quote(str(v))}" for k, v in params.items()])
            full_url = f"{url}?{query_string}"
            
            logger.info(f"访问 URL: {full_url}")
            
            # 使用 Selenium 加载页面
            self.driver.get(full_url)
            
            # 等待产品列表加载
            wait = WebDriverWait(self.driver, 15)
            try:
                # 等待产品容器加载
                wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, "[class*='tour-item'], [class*='item-tour'], li[data-itemid]")))
                logger.info("页面产品已加载")
            except:
                logger.warning("页面加载超时或找不到产品元素")
            
            self._random_delay(1, 2)
            
            # 获取页面源码
            html = self.driver.page_source
            soup = BeautifulSoup(html, 'html.parser')
            
            # 查找产品容器 - 尝试多种可能的选择器
            tour_items = []
            selectors = [
                soup.find_all('li', class_=lambda x: x and 'item-tour' in x),
                soup.find_all('div', class_=lambda x: x and 'tour-item' in x),
                soup.find_all('li', attrs={'data-itemid': True}),
                soup.find_all('a', class_=lambda x: x and 'tour-link' in x),
            ]
            
            for selector_result in selectors:
                if selector_result:
                    tour_items = selector_result
                    break
            
            logger.info(f"找到 {len(tour_items)} 个产品")
            
            if len(tour_items) == 0:
                logger.warning("未找到任何产品元素，可能需要调整选择器")
                return []
            
            page_tours = []
            for idx, item in enumerate(tour_items, 1):
                try:
                    tour = self._parse_tour_item(item)
                    if tour and self._is_valid_tour(tour):
                        page_tours.append(tour)
                        logger.debug(f"成功解析第 {idx} 个产品: {tour.get('产品名称', 'N/A')}")
                except Exception as e:
                    logger.debug(f"解析第 {idx} 个产品出错: {str(e)}")
                    continue
            
            logger.info(f"第 {page} 页成功解析了 {len(page_tours)} 个产品")
            return page_tours
            
        except Exception as e:
            logger.error(f"爬取页面出错: {str(e)}")
            return []
    
    def _parse_tour_item(self, item):
        """解析单个产品"""
        try:
            tour = {}
            
            # 产品名称 - 尝试多种方式获取
            name_elem = None
            for selector in [
                item.find('h3'),
                item.find('p', class_=lambda x: x and 'name' in x),
                item.find('a', class_=lambda x: x and 'title' in x),
                item.find('span', class_=lambda x: x and 'title' in x),
            ]:
                if selector:
                    name_elem = selector
                    break
            
            tour['产品名称'] = name_elem.text.strip() if name_elem else 'N/A'
            
            # 产品 URL
            link_elem = item.find('a', href=True)
            if link_elem:
                href = link_elem.get('href', '')
                tour['产品链接'] = href if href.startswith('http') else f"https://you.ctrip.com{href}"
            else:
                tour['产品链接'] = 'N/A'
            
            # 价格 - 尝试多种方式
            price = 'N/A'
            for price_selector in [
                item.find('span', class_=lambda x: x and 'price' in x),
                item.find('em', class_=lambda x: x and 'price' in x),
                item.find('i', class_=lambda x: x and 'rmb' in x),
            ]:
                if price_selector:
                    price_text = price_selector.text.strip()
                    price_match = re.search(r'\d+', price_text.replace(',', ''))
                    if price_match:
                        price = price_match.group()
                        break
            tour['价格'] = price
            
            # 天数
            days = 'N/A'
            for days_selector in [
                item.find('span', class_=lambda x: x and 'days' in x),
                item.find('span', class_=lambda x: x and 'day' in x),
                item.find('span', string=lambda s: s and '天' in s),
            ]:
                if days_selector:
                    days_text = days_selector.text.strip()
                    days_match = re.search(r'(\d+)\s*天', days_text)
                    if days_match:
                        days = days_match.group(1)
                        break
            tour['天数'] = days
            
            # 出发地
            tour['出发地'] = '多城市'  # 默认值
            
            # 目的地
            tour['目的地'] = self.destination
            
            # 评分
            score = 'N/A'
            for score_selector in [
                item.find('span', class_=lambda x: x and 'score' in x),
                item.find('span', class_=lambda x: x and 'rating' in x),
            ]:
                if score_selector:
                    score = score_selector.text.strip()
                    break
            tour['评分'] = score
            
            # 评价数
            review_count = '0'
            for review_selector in [
                item.find('span', class_=lambda x: x and 'review' in x),
                item.find('span', class_=lambda x: x and 'comment' in x),
                item.find('span', string=lambda s: s and '人' in s),
            ]:
                if review_selector:
                    review_text = review_selector.text.strip()
                    review_match = re.search(r'(\d+)', review_text)
                    if review_match:
                        review_count = review_match.group(1)
                        break
            tour['评价数'] = review_count
            
            # 关键词/特色 - 从产品名称中提取
            tour['关键词'] = self._extract_keywords(tour['产品名称'])
            
            return tour
            
        except Exception as e:
            logger.debug(f"解析产品出错: {str(e)}")
            return None
    
    @staticmethod
    def _extract_keywords(title):
        """从标题中提取关键词"""
        keywords = set()
        
        # 定义关键词映射
        keyword_map = {
            '温泉': ['温泉'],
            '漂流': ['漂流'],
            '自驾': ['自驾'],
            '跟团': ['跟团', '团队'],
            '蜜月': ['蜜月', '蜜月游'],
            '亲子': ['亲子', '亲子游'],
            '爸妈': ['爸妈', '父母'],
            '山水': ['山水', '名山', '大山'],
            '古镇': ['古镇', '水乡'],
            '海滨': ['海滨', '海边', '海岛', '沙滩'],
            '高铁': ['高铁'],
            '飞机': ['飞机'],
            '自由行': ['自由行', '自助游'],
            '周末': ['周末'],
            '短途': ['短途'],
            '长线': ['长线'],
            '国际': ['国际', '出国'],
            '港澳': ['港澳', '香港', '澳门'],
        }
        
        title_lower = title.lower()
        
        for key, patterns in keyword_map.items():
            for pattern in patterns:
                if pattern in title_lower:
                    keywords.add(key)
                    break
        
        return ','.join(keywords) if keywords else '其他'
    
    @staticmethod
    def _is_valid_tour(tour):
        """验证旅游产品数据"""
        required_fields = ['产品名称', '目的地']
        for field in required_fields:
            if not tour.get(field) or tour.get(field) == 'N/A':
                return False
        return True
    
    def save_to_excel(self, tours):
        """保存到 Excel 文件"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "旅游产品"
            
            # 定义列
            columns = ['产品名称', '价格', '出发地', '目的地', '天数', '评分', '评价数', '关键词', '产品链接']
            
            # 写入表头
            for col_idx, col_name in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = col_name
                
                # 设置表头样式
                cell.font = Font(bold=True, color="FFFFFF", size=12)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # 定义边框
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 写入数据
            for row_idx, tour in enumerate(tours, 2):
                for col_idx, col_name in enumerate(columns, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    value = tour.get(col_name, 'N/A')
                    cell.value = value
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    
                    # 价格列居右对齐
                    if col_name == '价格':
                        cell.alignment = Alignment(horizontal="right", vertical="center")
            
            # 调整列宽
            column_widths = {
                'A': 30,  # 产品名称
                'B': 12,  # 价格
                'C': 12,  # 出发地
                'D': 12,  # 目的地
                'E': 10,  # 天数
                'F': 10,  # 评分
                'G': 10,  # 评价数
                'H': 20,  # 关键词
                'I': 35,  # 产品链接
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # 设置行高
            ws.row_dimensions[1].height = 25
            for row in range(2, len(tours) + 2):
                ws.row_dimensions[row].height = 20
            
            # 冻结表头
            ws.freeze_panes = "A2"
            
            # 保存文件
            wb.save(self.output_file)
            logger.info(f"✅ Excel 文件已保存: {self.output_file}")
            logger.info(f"📊 共保存 {len(tours)} 个产品")
            
        except Exception as e:
            logger.error(f"保存 Excel 出错: {str(e)}")
            raise
    
    def close(self):
        """关闭浏览器"""
        if self.driver:
            try:
                self.driver.quit()
                logger.info("✅ 浏览器已关闭")
            except Exception as e:
                logger.warning(f"关闭浏览器出错: {str(e)}")
